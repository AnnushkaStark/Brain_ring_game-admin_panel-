import sqlite3 as sql
from openpyxl import load_workbook
connection = sql.connect('New_Brain_base')
cursor = connection.cursor()
workbook = load_workbook(filename = 'имя файла.xlsx')
worksheet = workbook.active
for row in worksheet.iter_rows(values_only= True):
    cursor.execute('INSERT INTO brain_base(вопрос, ответ, комментарий, уровень) VALUES(input(),input(),input(),int(input())',row)
connection.commit()
connection.close()