# ----------------------------------------------------------- #
# Данный файл содержит все необходимые функции для работы с   #
# базой данных.                                               #
# ----------------------------------------------------------- #

# -------------------- Импорт зависимостей ------------------ #
import sqlite3 as sql
from openpyxl import load_workbook
import logging as log

# ------------------------- Константы ----------------------- #
log.basicConfig(filename='brainring.log', level=log.DEBUG,
                format='%(asctime)s - %(levelname)s - %(message)s')
log.info("========== Файл database_controller.py ==========")

database = 'app/database/brainring.db'           # Имя БД
excel_file = 'app/database/questions_template.xlsx'    # Имя файла-шаблона, при загрузке через интерфейс - указывать путь выбранному файлу

# -------------------------- Запросы ------------------------ #
query_add_single_question = '''INSERT INTO questions (question, answer) VALUES (?, ?)'''     # Запрос на добавление записи в БД
query_get_all_questions = '''SELECT * FROM questions'''
query_get_single_question = ''' '''
query_delete_question = ''' '''
query_update_question = ''' '''
# -------------------------- ======= ------------------------ #
conn = None                         # Экземпляр соединения с БД

# -------------------------- Функции ------------------------ #

def get_connection():
    ''' Функция осуществляет подключение к базе данных database '''
    log.debug("==> get_connection() - вызвана функция")
    global conn
    if not conn:
        try:
            conn = sql.connect(database)
        except Exception as e:
            log.error("<== Не удалось подключиться к БД:", e)
            raise
        finally:
            log.debug("<== Соединение с БД создано")
    else:
        log.debug("<== Соединение с БД уже существует")
    return conn

    
    

# ----------------------------------------------------------- #

def get_all_questions():
    '''функция выводит все строки из базы данных'''
    log.debug("==> get_all_questions() - функция вызвана")
    try:
        with get_connection() as conn:
            result = conn.cursor().execute(query_get_all_questions).fetchall()
            log.debug(f"Вопросов из БД получено - {len(result)} шт.")
        log.debug("Связь с БД закрыта")
        log.debug("<== get_all_questions() - конец выполнения")
        return result
    except Exception as e:
        log.error("<== Не удалось получить вопросы из базы данных:", e)
    
# ----------------------------------------------------------- #
# ----------------------------------------------------------- #
# ----------------------------------------------------------- #

def add_questions_from_excel(file):
    ''' Добавление вопросов группой из Excel-файла. '''
    log.debug("==> add_questions_from_excel() - функция вызвана")
    # ДОБАВИТЬ: проверку расширения файла excel; продумать, на каком этапе её лучше делать

    try:    # Попытка открыть файл
        workbook = load_workbook(filename = file)  # Загружаем Excel-файл
        sheet = workbook['page']    # Указываем название страницы (можно ли переделать на индекс?)'

        count_added_questions = 0 # Счётчик добавленных вопросов
        with get_connection() as conn:
            cursor = conn.cursor()
            if sheet["A1"].value == "Вопрос" and sheet["B1"].value == "Ответ":      # Проверка на соответствие загружаемого файла шаблону
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    log.debug(f"Обрабатывается строка: {row}")
                    if row[0] == None:
                        log.warning("Поле 'Вопрос' не может быть пустым!")
                        continue
                    elif row[1] == None:
                        log.warning("Поле 'Ответ' не может быть пустым!")
                        continue
                    else:
                        try:
                            cursor.execute(query_add_single_question, (row[0], row[1]))
                        except sql.IntegrityError as sql_error:
                            log.warning(f"Вопрос уже существует в базе данных: {sql_error.args}")
                            continue
                        except Exception as e:
                            log.warning(f"Ошибка при добавлении вопроса: {e}")
                            continue
                    count_added_questions += 1
            else:
                log.error("<== Попытка использовать файл, не соответствующий шаблону!")
    except Exception as e:
        log.error("<== Невозможно открыть excel-файл:", e)
    finally:
        log.debug(f"<== add_question_from_excel() - конец выполнения. Добавлено {count_added_questions}/{sheet.max_row - 1} вопросов.") 
# ----------------------------------------------------------- #
# ----------------------------------------------------------- #
# ----------------------------------------------------------- #

def add_single_question(database: str, question: str, answer: str):
    ''' Добавление вопросов по одному по одному.'''

    connection = get_connection()   # Подключаемся к БД
    cursor = connection.cursor()
    question, answer = input("Введите вопрос: "), input("Введите ответ: ") # Строка временная, не нужна при работе через интерфейс

    try:
        if question != None and answer != None:
            cursor.execute(query_add_single_question, (question, answer))   # Непосредственно выполнение запроса к БД
            print('OK')
        else:
            raise sql.IntegrityError
    except sql.IntegrityError:
        print('Вопрос или ответ не может быть пустым!')
    except OSError as e:
        print(e)
        print('Такой вопрос уже существует!')
        print(f'Вопрос: {question}')
        print(f'Вопрос: {answer}')
        print()

    print('Уникальные вопросы добавлены')

    # Закрываем соединение с БД
    connection.commit()
    connection.close()

# def delete_question(question: str):
#     pass

def edit_question(question: str):
    pass





''' NEED REVIEW '''




# update_question = input()     #введите вопрос который хотите обновить
# update_answer = input() #введите ответ который хотите обновить
# conn = sql.connect('BrainRing.db')
#cursor = conn.cursor()

'''Функция выбирает вопрос для обновления из базы данных'''
def get_question(update_question):
    conn = sql.connect('BrainRing.db')
    cursor = conn.cursor()
    cursor.execute('SELECT вопрос FROM questions WHERE вопрос == update_question')
    result = cursor.fetchone()
    if result is not None:
        return result
    else:
        return 'Вопрос в базе не найден'
    conn.close

'''Функция выбирает ответ для обновления'''
def get_answer(update_answer):
    conn = sql.connect('BrainRing.db')
    cursor = conn.cursor()
    cursor.execute('SELECT ответ FROM questions WHERE ответ == update_answer')
    result = cursor.fetchone()
    conn.close()
    if result is not None:
        return result
    else:
        return 'Ответ в базе не найден'
    

new_question, new_answer  = '', ''

def update_question(update_question, update_answer):
    '''Эта функция обновляет вопрос '''
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE questions SET вопрос,  WHERE вопрос == update_question',(new_question))
        conn.commit()

def update_answer():
    '''Эта функция обновляет ответ'''

    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE questions SET answer,  WHERE answer == update_answer', (new_answer))
        conn.commit()
        cursor.close()

def delete_question(del_question):
    '''Эта функция удаляет строку с поиском по вопросу'''
    conn = sql.connect('BrainRing.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM questions WHERE вопрос LIKE del_question')  # Проверяем наличие строки  в таблице
    row = cursor.fetchone()
    if row is None:
        print('Вопрос не найден в базе данных')
    else:
        cursor.execute("DELETE FROM questions WHERE вопрос LIKE del_question")  # Удаляем строку
        conn.commit()
        print('Строка успешно удалена')
    cursor.close()
    conn.close()

def delete_answer(del_answer):
    '''Эта функция удаляет строку с поиском по ответу'''
    conn = sql.connect('BrainRing.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM questions WHERE ответ LIKE del_answer') #Проверяем наличие строки  в таблице
    row = cursor.fetchone()
    if row is None:
        print('Ответ не найден в базе данных')
    else:
        cursor.execute("DELETE FROM questions WHERE ответ LIKE del_answer") #Удаляем строку
        conn.commit()
        print('Строка успешно удалена')
    cursor.close()
    conn.close()


# name = input() #введите название базы данных

log.debug("===== Конец выполнения файла database_controller.py. =====")
