from typing import Self, TextIO
from PyQt6 import QtCore
from PyQt6.QtWidgets import *
from PyQt6.QtWidgets import QWidget
import sys
from PyQt6 import *
import sqlite3 as sql
from openpyxl import load_workbook
import logging as log
from dataface import *
from question_creator_gui import Creator

class Controller(Ui_MainWindow,QMainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.show()
        self.pushButton_connect.clicked.connect(self.get_connection)
        self.pushButton_load_exel.clicked.connect(self.add_questions_from_excel)
        self.lineEdit_new_question.text()
        self.lineEdit_new_answe.text()
        self.pushButton_append.clicked.connect(self.add_single_question)
        self.lineEdit_question_new.text()
        self.lineEdit_answer_new.text()
        self.lineEdit_question_old.text()
        self.lineEdit_answer_old.text()
        self.pushButton_Updae.clicked.connect(self.update_question)
        self.pushButton_select_all.clicked.connect(self.get_all_questions)
        self.lineEdit_delet_answer.text()
        self.pushButton_delete.clicked.connect(self.delete_question)
        self.pushButton_question_wibdow.clicked.connect(self.open_questions_creator)
        self.pushButton_Raiting_window.clicked.connect(self.open_raiting_window)
        self.count_added_questions = 0
        self.query_add_single_question = '''INSERT INTO questions (question, answer) VALUES (?, ?)'''     # Запрос на добавление записи в БД
        self.query_get_all_questions = '''SELECT * FROM questions'''
        self.query_get_single_question = '''SELECT question, answer FROM questions WHERE question LIKE ? OR answer LIKE ?'''
        self.query_update_question = '''UPDATE questions SET question = ? WHERE question LIKE ?'''
        self.query_update_answer = '''UPDATE questions SET answer = ? WHERE answer LIKE ?'''
        self.query_delete_question = '''DELETE FROM questions WHERE question LIKE ? OR answer LIKE ?'''
        self.foubd_question = ''
        self.database = 'app/assets/database/brainring.db'
        self.excel_file = 'app/assets/database/questions_template.xlsx'    # Имя файла-шаблона, при загрузке через интерфейс - указывать путь выбранному файлу
        self.conn = None

    def get_connection(self):
        ''' Функция осуществляет подключение к базе данных database '''
        log.debug("==> get_connection() - функция вызвана.\n")
        conn = sql.connect(self.database)
        if conn:
            try:
                result =  QMessageBox()
                result.setText('Cоединение создано')
                result.exec()
            except Exception as e:
                log.error(f"<== Не удалось подключиться к БД: {e}\n")
                raise
            finally:
                    log.debug("<== Соединение с БД создано.\n")

        else:        
            result =  QMessageBox()
            result.setText('Не удалось подключиться к базе')
            result.exec()
           
# ----------------------------------------------------------- #

    def get_all_questions(self):
        '''Функция выводит все вопросы и ответы в таблицу'''
        log.debug("==> get_all_questions() - функция вызвана.\n")
        try:
            with sql.connect(self.database ) as conn:
                result = conn.cursor().execute(self.query_get_all_questions).fetchall()
                self.tableWidget.setRowCount(len(result))
                self.tableWidget.setColumnCount(len(result[0]))
                log.debug(f"Вопросов из БД получено - {len(result)} шт.")
            log.debug("<== get_all_questions() - конец выполнения. Связь с БД закрыта.\n")   
            for row in range(len(result)):
                for column in range(len(result[row])):
                    item = QTableWidgetItem(str(result[row][column]))
                    self.tableWidget.setItem(row,column,item)
        except Exception as e:
            log.error(f"<== Не удалось получить вопросы из базы данных: {e}\n")

# ----------------------------------------------------------- #

    def get_single_question(self):
        ''' Функция возвращает из базы данных один вопрос, в котором есть соответствующий текст эта функция не привязана к виджетам'''
        log.debug("==> get_single_question() - функция вызвана.\n")
        with sql.connect(self.database) as conn:
            try:
                found_answer =self.lineEdit_answer_old() #Проверяем ввел ли пользователь данные
                found_question = self.lineEdit_question_old()
                result = conn.cursor().execute(self.query_get_all_questions).fetchall()
                log.debug("<== get_single_question() - конец выполнения.\n")
                for row in result:
                    if found_question == row[1] or found_answer == row[2]: #проверяем наличие вопроса в базе
                        print('Вопрос найден')
                else:
                    log.warning("<== Вопрос (ответ) не найден в БД.\n")
                    print('Вопрос не найден')
                    # WARNING - исправить на случай если вопросы не найдены, чтобы не было ошибки
            except Exception as e:
                log.warning(f"<== Не удалось получить вопрос: {e}\n")
        
# ----------------------------------------------------------- #

    def add_single_question(self):
        ''' Добавление одного вопроса'''
        log.debug("> add_single_question() - функция вызвана.\n")
        with sql.connect(self.database) as conn: 
            try:
                cursor = conn.cursor()
                new_question =  self.lineEdit_new_question.text()
                new_answer = self.lineEdit_new_answe.text()
                if new_question.strip() and new_answer.strip(): #Проверяем ввел ли пользоваетель данные
                    cursor.execute(self.query_add_single_question, (new_question, new_answer))
                    self.count_added_questions += 1
                    result =  QMessageBox()
                    result.setText('Вопрос добавлен')
                    result.exec()
                   
                else:
                    log.warning("< Поле  не может быть пустым!\n")
                    result =  QMessageBox()
                    result.setText('Поля не заполнены')
                    result.exec()
                    log.debug("< add_single_question() - конец выполнения.\n")
            except sql.IntegrityError as sql_error:
                log.warning(f"< Вопрос уже существует в базе данных: {sql_error.args}\n")
            except Exception as e:
                log.warning(f"< Ошибка при добавлении вопроса: {e}\n")
# ----------------------------------------------------------- #

    def add_questions_from_excel(self):
        ''' Добавление вопросов из Excel-файла. '''
        log.debug("==> add_questions_from_excel() - функция вызвана.\n")
        self.count_added_questions = 0
        # ДОБАВИТЬ: проверку расширения файла excel; продумать, на каком этапе её лучше делать

          # Попытка открыть файл
        try:
            workbook = load_workbook(self.excel_file)  # Загружаем Excel-файл
            sheet = workbook['page']    # Указываем название страницы (можно ли переделать на индекс?)'
           
            self.count_added_questions = 0 # Обнуление добавленных вопросов
            if sheet["A1"].value == "Вопрос" and sheet["B1"].value == "Ответ":
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    log.debug(f"Обрабатывается строка: {row}")
                    with sql.connect(self.database) as conn:
                        cursor = conn.cursor()
                        cursor.execute(self.query_add_single_question, (row[0], row[1]))
                result =  QMessageBox()
                result.setText('Вопросы добавлены')
                result.exec()
                log.debug(f"<== add_question_from_excel() - конец выполнения. Добавлено {self.count_added_questions}/{sheet.max_row - 1} вопросов.\n") 
            else:
                log.error("<== Попытка использовать файл, не соответствующий шаблону!\n")
                result =  QMessageBox()
                result.setText('Файл не соответствует формату')
                result.exec()
        except sql.IntegrityError as sql_error:
                log.warning(f"< Вопрос уже существует в базе данных: {sql_error.args}\n")
        except Exception as e:
            log.error(f"<== Невозможно открыть excel-файл: {e}\n")
                   
# ----------------------------------------------------------- #

    def update_question(self):
        ''' Функция изменяет текст вопроса и ответа для вопроса с идентификатором question_id'''
        log.debug("> update_question() - функция вызвана.\n")
        with sql.connect(self.database) as conn:
            try: # Пытаемся выполнить запрос по изменению вопроса в БД
                upd_question= self.lineEdit_question_new.text() #Проверяем ввел ли пользователь данные
                upd_answer =  self.lineEdit_answer_new.text()
                found_question = self.lineEdit_question_old.text()
                found_answer = self.lineEdit_answer_old.text()
                if upd_question.strip()  and upd_answer.strip() and found_question.strip() and found_answer.strip():
                    log.warning("< Поле  не может быть пустым!\n")
                    result = conn.cursor().execute(self.query_get_single_question,(found_question,found_answer))
                    if result:
                        conn.cursor().execute(self.query_update_question,(upd_question,found_question))
                        conn.cursor().execute(self.query_update_answer,(upd_answer,found_answer))
                        result =  QMessageBox()
                        result.setText('Вопрос\ответ изменен')
                        result.exec()
                        log.debug(f'< Вопрос  успешно изменён.\n')
                    else:
                        result =  QMessageBox()
                        result.setText('Ошибка при изменении вопроса')
                        result.exec()
                        
                else:
                    result =  QMessageBox()
                    result.setText('Изменяемый вопрос  не найден')
                    result.exec()
                   
            except Exception as e:
                log.warning(f"< Ошибка при изменении вопроса: {e}\n")

        
# ----------------------------------------------------------- #

    def delete_question(self):
        ''' Функция удаляет строку с поиском по вопросу или ответу'''
        log.debug("> delete_question() - функция вызвана.\n")
        with  sql.connect(self.database) as conn:
            try:
                del_question  = self.lineEdit_delet_question.text()
                del_answer = self.lineEdit_delet_answer.text()
                if del_question.strip() and del_answer.strip():
                    result = conn.cursor().execute(self.query_get_single_question,(del_question,del_answer))
                if result:
                    conn.cursor().execute(self.query_delete_question,(del_question,del_answer))
                    res = 'вопрос успешно удален'
                    log.debug(f'< Вопрос  успешно удалён.\n')
                else:
                    res = 'Заполнены не все поля или вопрос никогда не сущестовал'
                result =  QMessageBox()
                result.setText(res)
                result.exec()
            except Exception as e:
                log.error(f"< Ошибка при удалении вопроса: {e}\n")
        
# ----------------------------------------------------------- #
    def open_questions_creator(self):
        '''Эта функция открывает окно генерирующее вопросы'''
        self.question_creator = Creator()
        self.question_creator.show()
  


    def open_raiting_window(self):
        pass

log.debug("===== Конец выполнения файла database_controller.py. =====")


    
  
app = QApplication(sys.argv)
window = Controller()
sys.exit(app.exec())
