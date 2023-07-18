'''
Добавление вопросов из Excel-файла
'''

import sqlite3 as sql
from openpyxl import load_workbook

connection = sql.connect('brainring.db')
try:
    cursor = connection.cursor()
except:
    print('Connection error')
workbook = load_workbook(filename = 'questions_file.xlsx')
sheet = workbook['page']

query = '''INSERT INTO questions(question, answer)
           VALUES(?,?)'''

if sheet.row[0][0] == "Вопрос" and sheet.row[0][1] == "Ответ": # Не работает
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            cursor.execute(query, (row[0], row[1]))
            print('OK')
        except:
            print('Такой вопрос уже существует!')
            print(f'Вопрос: {row[0]}')
            print(f'Вопрос: {row[1]}')
            print()
else:
    print("Используйте только шаблон предоставленный программой!")

print('Уникальные вопросы добавлены')
connection.commit()
connection.close()